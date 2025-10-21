import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook



def readExcel(sheet,max_row,max_col):
    for row in sheet.iter_rows(max_row=max_row,
                               max_col=max_col):
        for cell in row:
            print(cell.value, end='\t')
            print()



def create_excel_file(file_path,data):
    wb=Workbook()
    ws=wb.active
    ws.title="sample Data"

    for row in data:
        ws.append(row)
    wb.save(file_path)
    print("file save complete")
    return ws


if __name__=="__main__":
    data_list=[
        ["이름","나이","국가"],
        ["홍길동",30,"대한민국"],
        ["Alice",25,"USA"],
        ["Bob",28,"UK"]
    ]
    excel_path=r"ch22\excel\new_example.xlsx"

    sheet=create_excel_file(excel_path,data_list)
    readExcel(sheet,3,5)