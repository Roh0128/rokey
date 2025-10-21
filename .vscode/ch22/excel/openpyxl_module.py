import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# book>sheet>cell

wb =Workbook()
sheet=wb.active

sheet['A1']='Hello'
sheet['B1']='World'


file=r"ch22\excel\new_example.xlsx"
wb.save(file)
# wb=load_workbook("")
# sheet=wb.active()


wb=load_workbook(file)
sheet= wb.active

print(sheet['A1'].value)
print(sheet['B1'].value)

print("-----------------")
# 여러 셀의 데이터 읽기

for row in sheet.iter_rows(min_row=1, max_row=5, max_col=3):
    for cell in row:
        print(cell.value)
print("------------------------------")
def readExcel(max_row, max_col):
    for row in sheet.iter_rows(max_row=max_row,max_col=max_col):
        for cell in row:
            print(cell.value, end="\t")
            print()
readExcel(5,3)

# 셀 값 수정하기
sheet["A1"]= '새로운 값'

file=r"ch22\excel\new_example.xlsx"
wb.save(file)

readExcel(2,3)

print("--------------------")

# 다양한 데이터 형식 다루기
# 숫자, 날짜, 수식 입력(실험결과는 엑설을 열어 확인)

from openpyxl.utils import get_column_letter
from datetime import datetime
wb=Workbook()
sheet=wb.active

sheet['A1']=datetime(2025,10,15)

sheet['B1']=100

sheet['C1']= '=B1*2'

file=r"ch22\excel\formulas_example.xlsx"
wb.save(file)

readExcel(3,3)